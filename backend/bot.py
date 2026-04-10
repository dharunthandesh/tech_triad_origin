import os
import json
import google.generativeai as genai
from dotenv import load_dotenv
from supabase import create_client, Client
import openpyxl
from datetime import datetime

load_dotenv()

# ─── Configure Gemini ─────────────────────────────────────────────────────────
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
if not GEMINI_API_KEY:
    print("[WARN] GEMINI_API_KEY not found in environment!")
genai.configure(api_key=GEMINI_API_KEY)

# ─── Supabase Client ──────────────────────────────────────────────────────────
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")
supabase: Client = None
if SUPABASE_URL and SUPABASE_KEY:
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
else:
    print("[WARN] Supabase credentials missing!")

# ─── Excel Setup ──────────────────────────────────────────────────────────────
EXCEL_FILE = "bookings.xlsx"
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bookings"
        ws.append(["S.No", "Name", "Service", "Appointment Time", "Phone / ID", "Channel", "Booking Timestamp"])
        wb.save(EXCEL_FILE)

init_excel()

def save_to_excel(name, service, time, phone, channel):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        s_no = ws.max_row
        ws.append([s_no, name, service, time, phone, channel, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        wb.save(EXCEL_FILE)
        print(f"[SUCCESS] Booking saved to Excel: {name} - {service} at {time}")
        return True
    except Exception as e:
        print(f"[ERROR] Error saving to Excel: {e}")
        return False

def save_to_supabase(name, service, time, phone, source="whatsapp"):
    if not supabase:
        print("[WARN] Supabase not configured, skipping.")
        return False
    try:
        supabase.table("leads").insert({
            "phone": phone,
            "name": name,
            "service": service,
            "booking_time": time,
            "status": "new",
            "source": source
        }).execute()
        print(f"[SUCCESS] Lead saved to Supabase: {name} - {service}")
        return True
    except Exception as e:
        print(f"[ERROR] Error saving to Supabase: {e}")
        return False

def save_booking(name, service, time, phone, channel="Voice"):
    save_to_excel(name, service, time, phone, channel)
    save_to_supabase(name, service, time, phone, channel)

# ─── System Prompt ────────────────────────────────────────────────────────────
SYSTEM_INSTRUCTION = """
You are a professional and friendly virtual receptionist for a hair salon named "Glamour Studio".

Your goal is to assist customers in booking appointments via WhatsApp or Voice Call.

Salon Details:
- Services: Haircut (300 rupees), Facial (800 rupees), Hair Spa (1200 rupees), Bridal Makeup (5000 rupees)
- Timings: 10 AM to 8 PM, all days

Booking Flow:
1. Greet the customer warmly.
2. Ask what service they need.
3. Ask for their name.
4. Ask for their preferred date and time.
5. Confirm the full booking and thank them.

CRITICAL RULE:
Once the booking is CONFIRMED (you have name + service + time), you MUST include this at the very end of your reply:
DATA:{"name": "NAME", "service": "SERVICE", "time": "TIME"}

Example:
"Perfect! I've booked a Haircut for you at 5 PM on Tuesday, Ravi. See you then! DATA:{"name": "Ravi", "service": "Haircut", "time": "5 PM Tuesday"}"

Important Rules:
- NEVER say you are an AI or a bot. Always act as a human receptionist.
- For Voice calls: Keep responses SHORT (1-2 sentences max). No emojis. Speak naturally.
- For WhatsApp: Be warm and friendly, use emojis where appropriate.
"""

# ─── In-memory session storage ────────────────────────────────────────────────
# We store genai.ChatSession objects here, keyed by sender_id
user_chats = {}

def get_ai_response(user_message, sender_id, channel="Voice"):
    """
    Get AI response. Maintains per-caller conversation history using Gemini ChatSession.
    channel: 'Voice' or 'WhatsApp'
    """
    if sender_id not in user_chats:
        sys_prompt = SYSTEM_INSTRUCTION
        if channel == "Voice":
            sys_prompt += "\nIMPORTANT: This is a LIVE VOICE CALL. Keep ALL responses under 2 short sentences. No emojis. Speak naturally like a human receptionist."
        
        # Initialize standard gemini model with system instruction
        model = genai.GenerativeModel(
            model_name="gemini-2.5-flash",
            system_instruction=sys_prompt,
        )
        # Start a chat session (maintains history automatically)
        user_chats[sender_id] = model.start_chat(history=[])

    chat = user_chats[sender_id]

    try:
        response = chat.send_message(user_message)
        text = response.text

        # Check for booking confirmation
        if "DATA:{" in text:
            parts = text.split("DATA:")
            clean_text = parts[0].strip()
            json_str = parts[1].strip()
            try:
                booking_data = json.loads(json_str)
                save_booking(
                    booking_data.get("name"),
                    booking_data.get("service"),
                    booking_data.get("time"),
                    sender_id,
                    channel
                )
            except Exception as e:
                print(f"[WARN] Could not parse booking JSON: {e} | Raw: {json_str}")
            return clean_text

        return text

    except Exception as e:
        error_msg = f"Gemini API Error: {str(e)[:150]}"
        print(f"[ERROR] {error_msg}")
        with open("debug_log.txt", "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now()}] {error_msg}\n")
        if channel == "Voice":
            return "I'm sorry, I had a small technical issue. Could you please repeat that?"
        return "I'm sorry, I'm having a little trouble right now! Please try again in a moment."
